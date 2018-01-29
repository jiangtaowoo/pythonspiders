# -*- coding: utf-8 -*-
import sys
from sys import argv
import os
import datetime
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment
from baseadaptor import AdaptorSqlite


def add_worksheet_headers(dbrow, wb, shname, col_name_list, start_idx):
    ws = wb.active
    if 'heet' in ws.title:
        ws.title = shname
    else:
        ws = wb.create_sheet(shname)
    al = Alignment(horizontal="center", vertical="center")
    for colidx, colname in enumerate(col_name_list):
        if colidx >= start_idx:
            curcol = colidx-start_idx+1
            first_cell = ws.cell(row=1, column=curcol)
            first_cell.alignment = al
            ws.cell(row=1, column=curcol, value=colname)
            #ws.merge_cells(start_row=1, start_column=curcol, end_row=2, end_column=curcol)
    return ws

def adjust_column_width(wb):
    for shname in wb.sheetnames:
        ws = wb[shname]
        if 'jd.com' in shname:
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['G'].width = 16
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['K'].width = 50
            ws.column_dimensions['N'].width = 20
        else:
            ws.column_dimensions['A'].width = 14
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['G'].width = 16
            ws.column_dimensions['H'].width = 50
            ws.column_dimensions['I'].width = 16

def output_product_info(sqliteada, wb, xlsmodelcfgs, id_row_mapping, colcnt_mapping, tbname, cond_dict=None):
    # step1. get product details and output
    #id_row_mapping = {}  # {'sheet1':{'productid1':row1, 'productid2':row2}}
    #colcnt_mapping = {}
    for website, modelcfg in xlsmodelcfgs.iteritems():
        if tbname in modelcfg:
            col_list = modelcfg[tbname]
            colcnt_for_details = len(col_list)-2
            #cond = {'website': website}
            if cond_dict:
                col_name_list, data_set = sqliteada.load_data(tbname, col_list, **cond_dict)
            else:
                col_name_list, data_set = sqliteada.load_data(tbname, col_list)#**cond)
            if not data_set:
                continue
            ws = None
            start_idx = 2
            for row in data_set:
                shname = row[0]
                if shname not in wb.sheetnames:
                    ws = add_worksheet_headers(row, wb, shname, col_name_list, start_idx)
                    id_row_mapping[shname] = {}
                    colcnt_mapping[shname] = colcnt_for_details
                else:
                    ws = wb[shname]
                rowmapping = id_row_mapping[shname]
                rowidx = len(rowmapping) + 2
                rowmapping[row[1]] = rowidx
                for colidx, colname in enumerate(col_name_list):
                    if colidx >= start_idx:
                        ws.cell(row=rowidx, column=colidx - start_idx + 1, value=row[colidx])
    #return id_row_mapping, colcnt_mapping

def get_sheetname_by_productid(prodid, id_row_mapping):
    for shname, rowmapping in id_row_mapping.iteritems():
        if prodid in rowmapping:
            return shname, rowmapping[prodid]
    return None, None

def output_ext_info(sqliteada, wb, xlsmodelcfgs, id_row_mapping, colcnt_mapping, tbname, cond_dict=None):
    start_col = 1
    for website, modelcfg in xlsmodelcfgs.iteritems():
        if tbname in modelcfg:
            col_list = modelcfg[tbname]
            cond = {'website': website}
            if cond_dict:
                col_name_list, data_set = sqliteada.load_data(tbname, col_list, **cond_dict)
            else:
                col_name_list, data_set = sqliteada.load_data(tbname, col_list)#**cond)
            if not data_set:
                continue
            col_step = 1  # price and viprice occupy two columns
            #if len(col_list)>4:
            #    col_step = 2
            ws = None
            al = Alignment(horizontal="center", vertical="center")
            col_updated_dict = dict()
            #date0 = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            for row in data_set:
                prodid = row[0]
                shname, rowidx = get_sheetname_by_productid(prodid, id_row_mapping)
                #shname = row[0]
                if shname not in col_updated_dict:
                    start_col = colcnt_mapping[shname]
                    colcnt_mapping[shname] = start_col + len(col_list)
                    start_idx = start_col + col_step * 1#(date1 - date0).days
                    ws = wb[shname]
                    #headers
                    for colidx, colname in enumerate(col_name_list):
                        if colidx > 0:
                            ws.cell(row=1, column=start_idx+colidx, value=col_name_list[colidx])
                    col_updated_dict[shname] = True
                #rowmapping = id_row_mapping[shname]
                #rowidx = rowmapping[internal_id]
                #first_cell = ws.cell(row=1, column=start_idx)
                #first_cell.alignment = al
                for colidx, colname in enumerate(col_name_list):
                    if colidx > 0:
                        ws.cell(row=rowidx, column=start_idx+colidx, value=row[colidx])

def sqlite_to_xlsx(xls_file_name):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    db_file_path = os.path.sep.join([current_dir, 'products.db'])
    id_row_mapping = {}  # {'sheet1':{'prodid1':row1, 'prodid2':row2}}
    colcnt_mapping = {}
    wb = Workbook()
    sqliteada = AdaptorSqlite(spidername='', dbname='comments.db')
    sqliteada.load_db_config(os.path.sep.join([current_dir, 'config', 'dbs.yaml']))
    xlsmodelcfgs = yaml.load(open(os.path.sep.join([current_dir, 'config', 'xlsmodel.yaml'])))
    output_product_info(sqliteada, wb, xlsmodelcfgs, id_row_mapping, colcnt_mapping, 'tmall_comment')
    output_product_info(sqliteada, wb, xlsmodelcfgs, id_row_mapping, colcnt_mapping, 'jd_comment')
    #output_ext_info(sqliteada2, wb, xlsmodelcfgs, id_row_mapping, colcnt_mapping, 'prod_addinfo')
    adjust_column_width(wb)
    wb.save(xls_file_name)

if __name__=="__main__":
    script, xls_file_name = argv
    if '.xlsx' in xls_file_name:
        sqlite_to_xlsx(xls_file_name)
        print 'finish!'
#    sqlite_to_xlsx('products.db', ''.join(['products_', str(datetime.date.today()), '.xlsx']), '2017-12-7')