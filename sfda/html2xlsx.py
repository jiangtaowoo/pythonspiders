# -*- coding: utf-8 -*-
import os
from sys import argv
import lxml.html
from openpyxl import Workbook
from openpyxl.styles import Alignment

def get_prod_filename(themonth, id, datatype=None):
    if datatype is None:
        return ''.join(['.', os.sep, 'dataout', os.sep, str(themonth), os.sep, str(id), '.html'])
    else:
        return ''.join(['.', os.sep, 'dataout', os.sep, 'sfda_', str(datatype), '_', str(themonth), '.log'])

def read_product_html(prod_info_file_path):
    if os.path.exists(prod_info_file_path):
        infile = open(prod_info_file_path)
        return lxml.html.fromstring(infile.read())
    return None

def output_xpath_text(ws, cur_row, cur_col, html, xpath_expression):
    content_list = html.xpath(xpath_expression)
    if len(content_list)>0:
        ws.merge_cells(start_row=cur_row, start_column=cur_col, end_row=cur_row, end_column=cur_col+2)
        ws.cell(row=cur_row, column=cur_col, value=_stringify_node(content_list[0]))
        return cur_row+1
    return cur_row

def _stringify_node(node):
    parts = [x for x in node.itertext()]
    return ''.join(filter(None, parts)).strip()

def output_one_product_info(xlsxfilename, wb, cur_row, themonth, file_idx, prod_info_file_name, prod_page_idx=0):
    prod_info_file_path = '.' + os.sep + 'dataout' + os.sep + str(themonth) + os.sep + prod_info_file_name
    #creating worksheet
    ws = wb.active
    shname = 'J2017' + str(themonth)
    if 'heet' in ws.title:
        ws.title = shname
        wsidx = wb.create_sheet(shname + '_INDEX')
    else:
        ws = wb[shname] #.create_sheet(shname)
        wsidx = wb[shname + '_INDEX']
    wsidx.cell(row=file_idx+1, column=1, value=prod_info_file_name)
    fcell = wsidx.cell(row=file_idx+1, column=1)
    fcell.hyperlink = prod_info_file_path
    #wsidx.cell(row=file_idx+1, column=2, value=cur_row)
    fcell = wsidx.cell(row=file_idx+1, column=2)
    cvalue = str(cur_row)
    hlink = ''.join(['#', shname, '!A', cvalue])
    fcell.value = '=HYPERLINK("%s", "%s")' % (hlink, cvalue)
    wsidx.cell(row=file_idx+1, column=3, value=prod_page_idx)
    wsidx.cell(row=file_idx+1, column=4, value='=%s!A%s' % (shname,cvalue))
    #getting data
    xpath_title_zh = '//table[@class="pageEnd"]/tbody/tr[5]/td'
    xpath_title_py = '//table[@class="pageEnd"]/tbody/tr[6]/td'
    xpath_peifang_title = '//table[@class="pageEnd"]/tbody/tr[8]/td/table/tbody/tr[@height]/td'
    xpath_peifang_detail = '//table[@class="pageEnd"]/tbody/tr[8]/td/table/tbody/tr[not(@height)]/td'
    xpath_peifang = './table/tbody//tr'
    html = read_product_html(prod_info_file_path)
    if html is None:
        return cur_row
    #output titles
    cur_row = output_xpath_text(ws, cur_row, 1, html, xpath_title_zh)
    cur_row = output_xpath_text(ws, cur_row, 1, html, xpath_title_py)
    #output tables
    rows_title = html.xpath(xpath_peifang_title)
    rows_detail = html.xpath(xpath_peifang_detail)
    for pf_idx, pf_title in enumerate(rows_title):
        #表头
        ws.cell(row=cur_row, column=2, value=_stringify_node(pf_title))
        cur_row += 1
        #表内容
        rows = rows_detail[pf_idx].xpath(xpath_peifang)
        #rows = html.xpath(xpath_peifang)
        merge_ind_list = [[0]*3 for i in range(len(rows))]
        def find_start_col(merge_ind_list_row, from_idx):
            for idx, x in enumerate(merge_ind_list_row):
                if idx<from_idx:
                    continue
                if x==0:
                    return idx
            return -1
        alcc = Alignment(horizontal="center", vertical="center")
        alvc = Alignment(vertical="center")
        alhc = Alignment(horizontal="center")
        for idx, rowhtml in enumerate(rows):
            cols = rowhtml.xpath('.//td')
            from_idx = 0
            for colhtml in cols:
                col_idx = find_start_col(merge_ind_list[idx], from_idx)
                first_cell = ws.cell(row=cur_row, column=col_idx+1)
                if idx==0 or col_idx==0:
                    first_cell.alignment = alhc
                #mark rowspan information
                if 'rowspan' in colhtml.attrib:
                    rowspan = int(colhtml.attrib['rowspan'])
                    if rowspan>1:
                        for j in xrange(1,rowspan):
                            merge_ind_list[idx+j][col_idx] = 1
                        if col_idx==0:
                            first_cell.alignment = alcc
                        else:
                            first_cell.alignment = alvc
                        ws.merge_cells(start_row=cur_row, start_column=col_idx+1, end_row=cur_row+rowspan-1, end_column=col_idx+1)
                #output data
                ws.cell(row=cur_row, column=col_idx+1, value=_stringify_node(colhtml))
                from_idx = col_idx + 1
            cur_row += 1
    cur_row += 1
    return cur_row

def main(themonth):
    orderf_d = dict()
    with open(get_prod_filename(themonth, 0, "order")) as orderf:
        for idx, line in enumerate(orderf):
            line = line.strip().split('\t')
            orderf_d[line[0]+'.html'] = (idx, line[1])
    flist = os.listdir('.' + os.sep + 'dataout' + os.sep + str(themonth))
    flist.sort(key=lambda x: orderf_d[x][0] if x in orderf_d else -1)
    wb = Workbook()
    cur_row = 1
    flen = len(flist)
    xlsxfilename = '.' + os.sep + 'dataout' + os.sep + str(themonth) + '.xlsx'
    for idx, f in enumerate(flist):
        page_idx = orderf_d[f][1]
        cur_row = output_one_product_info(xlsxfilename, wb, cur_row, themonth, idx, f, page_idx)
        print ' - %s  %d of %d' % (f, idx+1, flen)
    adjust_column_width(wb)
    wb.save(xlsxfilename)

def adjust_column_width(wb):
    for shname in wb.sheetnames:
        ws = wb[shname]
        if 'INDEX' in shname:
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 60
        else:
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 20

if __name__=="__main__":
    if len(argv)==2:
        script, themonth = argv
        main(themonth)
    else:
        for i in xrange(0,10):
            print '>>> cata %d ---' % (i)
            main(i)