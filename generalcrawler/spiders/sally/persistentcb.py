# -*- coding: utf-8 -*-
import os
import datetime
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment
from persistadaptor.baseadaptor import AdaptorSqlite


def add_worksheet_headers(dbrow, wb, shname, col_name_list, start_idx):
    ws = wb.active
    if 'heet' in ws.title:
        ws.title = shname
    else:
        ws = wb.create_sheet(shname)
    al = Alignment(horizontal="center", vertical="center")
    for colidx, colname in enumerate(col_name_list):
        if colidx >= start_idx:
            curcol = colidx-start_idx+1
            first_cell = ws.cell(row=1, column=curcol)
            first_cell.alignment = al
            ws.cell(row=1, column=curcol, value=colname)
            ws.merge_cells(start_row=1, start_column=curcol, end_row=2, end_column=curcol)
    return ws


def adjust_column_width(wb):
    for shname in wb.sheetnames:
        ws = wb[shname]
        for col in ws.columns:
            max_length = 0
            column = col[0].column  # Get the column name
            for cell in col:
                if cell.row>1:
                    if cell.value:
                        if 'http' not in cell.value:
                            if isinstance(cell.value, unicode):
                                if len(cell.value)*1.2 > max_length:
                                    max_length = len(cell.value)*1.2
                            else:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                        else:
                            max_length = 10
            adjusted_width = (max_length + 2) * 1.2
            adjusted_width = adjusted_width if adjusted_width<60 else 60
            ws.column_dimensions[column].width = adjusted_width

def output_product_info(sqliteada, wb, xlsmodelcfgs):
    # step1. get product details and output
    tenant_row_mapping = {}  # {'sheet1':{'productid1':row1, 'productid2':row2}}
    colcnt_mapping = {}
    for website, modelcfg in xlsmodelcfgs.iteritems():
        tbname = 'prod_details'
        if tbname in modelcfg:
            col_list = modelcfg[tbname]
            colcnt_for_details = len(col_list)
            cond = {'website': website}
            col_name_list, data_set = sqliteada.load_data(tbname, col_list, **cond)
            if not data_set:
                continue
            ws = None
            start_idx = 2
            for row in data_set:
                shname = row[0]
                if shname not in wb.sheetnames:
                    ws = add_worksheet_headers(row, wb, shname, col_name_list, start_idx)
                    tenant_row_mapping[shname] = {}
                    colcnt_mapping[shname] = colcnt_for_details
                else:
                    ws = wb[shname]
                rowmapping = tenant_row_mapping[shname]
                rowidx = len(rowmapping) + 3
                rowmapping[row[1]] = rowidx
                for colidx, colname in enumerate(col_name_list):
                    if colidx >= start_idx:
                        ws.cell(row=rowidx, column=colidx - start_idx + 1, value=row[colidx])
    return tenant_row_mapping, colcnt_mapping

def output_price_info(sqliteada, wb, start_date, tenant_row_mapping, colcnt_mapping, xlsmodelcfgs):
    start_col = 1
    for website, modelcfg in xlsmodelcfgs.iteritems():
        if 'prod_price' in modelcfg:
            col_list = modelcfg['prod_price']
            cond = {'website': website}
            col_name_list, data_set = sqliteada.load_data('prod_price', col_list, **cond)
            if not data_set:
                continue
            col_step = 1  # price and viprice occupy two columns
            if len(col_list)>4:
                col_step = 2
            ws = None
            al = Alignment(horizontal="center", vertical="center")
            date0 = datetime.datetime.strptime(start_date, "%Y-%m-%d")
            for row in data_set:
                shname = row[0]
                start_col = colcnt_mapping[shname]
                internal_id = row[1]
                date1 = datetime.datetime.strptime(row[2], "%Y-%m-%d")
                ws = wb[shname]
                rowmapping = tenant_row_mapping[shname]
                rowidx = rowmapping[internal_id]
                colidx = start_col + col_step * (date1 - date0).days
                first_cell = ws.cell(row=1, column=colidx)
                first_cell.alignment = al
                ws.cell(row=1, column=colidx, value=row[2])
                ws.cell(row=2, column=colidx, value=col_name_list[3])
                ws.cell(row=rowidx, column=colidx, value=row[3])
                if len(col_list)>4:
                    ws.merge_cells(start_row=1, start_column=colidx, end_row=1, end_column=colidx+1)
                    ws.cell(row=2, column=colidx+1, value=col_name_list[4])
                    ws.cell(row=rowidx, column=colidx + 1, value=row[4])

def sqlite_to_xlsx(xls_file_name, start_date):
    current_dir = os.path.dirname(os.path.abspath(__file__))
    db_file_path = os.path.sep.join([current_dir, 'products.db'])
    wb = Workbook()
    sqliteada = AdaptorSqlite(spidername='sally')
    sqliteada.load_db_config(os.path.sep.join([current_dir, 'config', 'dbs.yaml']))
    xlsmodelcfgs = yaml.load(open(os.path.sep.join([current_dir, 'config', 'xlsmodel.yaml'])))
    tenant_row_mapping, colcnt_mapping = output_product_info(sqliteada, wb, xlsmodelcfgs)
    output_price_info(sqliteada, wb, start_date, tenant_row_mapping, colcnt_mapping, xlsmodelcfgs)
    adjust_column_width(wb)
    wb.save(xls_file_name)

#if __name__=="__main__":
#    sqlite_to_xlsx('products.db', ''.join(['products_', str(datetime.date.today()), '.xlsx']), '2017-12-7')